import openpyxl
import msvcrt
import json
import math
from sklearn.mixture import GaussianMixture
from ultralytics import YOLO
import cv2
import random
from camera import Camera
import numpy as np
from os.path import join
import os
os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"




def list_dir(path, list_name, extension):
    for file in os.listdir(path):
        file_path = os.path.join(path, file)
        if os.path.isdir(file_path):
            list_dir(file_path, list_name, extension)
        else:
            if file_path.endswith(extension):
                list_name.append(file_path)
    try:
        list_name = sorted(list_name, key=lambda k: int(
            os.path.split(k)[1].split(extension)[0].split('_')[-1]))
    except Exception as e:
        print(e)
    return list_name


def overlay(image, mask, color, alpha, resize=None):
    """Combines image and its segmentation mask into a single image.

    Params:
        image: Training image. np.ndarray,
        mask: Segmentation mask. np.ndarray,
        color: Color for segmentation mask rendering.  tuple[int, int, int] = (255, 0, 0)
        alpha: Segmentation mask's transparency. float = 0.5,
        resize: If provided, both image and its mask are resized before blending them together.
        tuple[int, int] = (1024, 1024))

    Returns:
        image_combined: The combined image. np.ndarray

    """
    # color = color[::-1]
    colored_mask = np.expand_dims(mask, 0).repeat(3, axis=0)
    colored_mask = np.moveaxis(colored_mask, 0, -1)
    masked = np.ma.MaskedArray(image, mask=colored_mask, fill_value=color)
    image_overlay = masked.filled()

    if resize is not None:
        image = cv2.resize(image.transpose(1, 2, 0), resize)
        image_overlay = cv2.resize(image_overlay.transpose(1, 2, 0), resize)

    image_combined = cv2.addWeighted(image, 1 - alpha, image_overlay, alpha, 0)
    return image_combined


def save_data(object_records):
    # Save data to a file
    with open("data6.json", "w") as file:
        json.dump(object_records, file)
        print("success")


def plot_one_box(x, img, color=None, label=None, line_thickness=2):
    # Plots one bounding box on image img
    tl = line_thickness or round(
        0.002 * (img.shape[0] + img.shape[1]) / 2) + 1  # line/font thickness
    color = color or [random.randint(0, 255) for _ in range(3)]
    c1, c2 = (int(x[0]), int(x[1])), (int(x[2]), int(x[3]))
    cv2.rectangle(img, c1, c2, color, thickness=tl, lineType=cv2.LINE_AA)
    if label:
        tf = max(tl - 1, 1)  # font thickness
        t_size = cv2.getTextSize(label, 0, fontScale=tl / 3, thickness=tf)[0]
        c2 = c1[0] + t_size[0], c1[1] - t_size[1] - 3
        cv2.rectangle(img, c1, c2, color, -1, cv2.LINE_AA)  # filled
        cv2.putText(img, label, (c1[0], c1[1] - 2), 0, tl / 3,
                    [225, 255, 255], thickness=tf, lineType=cv2.LINE_AA)


def calculate_iou(bbox1, bbox2):
    # bbox为(x_min, y_min, x_max, y_max)格式的BBOX
    x1, y1, x2, y2 = bbox1
    x3, y3, x4, y4 = bbox2

    # 计算BBOX1和BBOX2的交集区域
    x_left = max(x1, x3)
    y_top = max(y1, y3)
    x_right = min(x2, x4)
    y_bottom = min(y2, y4)

    intersection_area = max(0, x_right - x_left + 1) * \
        max(0, y_bottom - y_top + 1)

    # 计算BBOX1和BBOX2的并集区域
    bbox1_area = (x2 - x1 + 1) * (y2 - y1 + 1)
    bbox2_area = (x4 - x3 + 1) * (y4 - y3 + 1)
    union_area = bbox1_area + bbox2_area - intersection_area

    # 计算IoU值
    iou = intersection_area / union_area

    return iou


def estimate_length_clusters(lengths, min_components=1, max_components=5):
    # Reshape the data into a column vector
    X = lengths.reshape(-1, 1)

    # Determine the optimal number of clusters
    lowest_bic = np.infty
    best_gmm = None
    best_n_components = None

    for n_components in range(min_components, max_components+1):
        gmm = GaussianMixture(n_components=n_components)
        gmm.fit(X)
        bic = gmm.bic(X)
        if bic < lowest_bic:
            lowest_bic = bic
            best_gmm = gmm
            best_n_components = n_components

    # Get the estimated means of the Gaussian components
    estimated_means = best_gmm.means_

    return estimated_means


def combine_measurements(measurements, confidences):
    grades = ["Grade1", "Grade2", "Grade3"]

    # Initialize the total confidence and the weighted sum for each grade
    total_confidence = 0.0
    weighted_sum = {grade: 0.0 for grade in grades}

    # Iterate over each measurement and confidence level
    for measurement, confidence in zip(measurements, confidences):
        # If grade3 is detected with high confidence, return grade3 immediately
        if measurement == "Grade3" and confidence >= 0.9:
            return "Grade3"

        # Update the weighted sum for the corresponding grade
        weighted_sum[measurement] += confidence
        total_confidence += confidence

    # Calculate the weighted average for each grade
    weighted_average = {
        grade: weighted_sum[grade] / total_confidence for grade in grades}

    # Determine the final result based on the highest weighted average
    final_result = max(weighted_average, key=weighted_average.get)
    return final_result

def GetFinalGrade(Final_list):

    id, m_len, mid_len, m_wid, mid_wid, f_grad=Final_list
    if mid_len<9 and mid_len>=6 and mid_wid <3.5 and (f_grad=="Normal" or f_grad=="Minor defects"):
        F_SampleGrade="Premium"
    elif mid_len<6 and mid_len>=3 and mid_wid <3.5 and (f_grad=="Normal" or f_grad=="Minor defects"):
        F_SampleGrade = "Good"
    elif mid_wid <3.5 and f_grad=="Severe defects":
        F_SampleGrade = "Fair"
    else:
        F_SampleGrade = "Cull"
    return F_SampleGrade


def FindClassAndCof(box1, box2s):
    # box1 from track
    # box2s all boxes from grade
    thr = 100
    for idx, box2 in enumerate(box2s):
        box2 = box2[0:4].tolist()
        iou = calculate_iou(box1, box2)
        if iou < thr:
            thr = iou
            final_idx = idx
    if 'final_idx' in locals():
        pass
    else:
        final_idx = 1
        print("final_idx 没有被赋值或绑定")
    return final_idx


def append_to_xlsx(data_list, file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active
    row = ws.max_row + 1

    for i, item in enumerate(data_list, 1):
        ws.cell(row=row, column=i, value=item)

    wb.save(file_path)


def pipeline(input_type='camera', source=None, model_paths=None, save_dir=None):
    # Load a model
    result_dict = {}
    model_track = YOLO(model_paths[0])
    model_grade = YOLO(model_paths[1])
    class_names = model_grade.names
    print('Class Names: ', class_names)
    colors = [[random.randint(0, 255) for _ in range(3)] for _ in class_names]
    colors = [[0, 191, 255], [205, 255, 121], [180, 105, 255]]

    if input_type == 'camera':
        camera = source
    elif input_type == 'video':
        pass
    elif input_type == 'image':
        im_paths = source

    offset = 0.0330  # in/px
    line_position = 640 - 129
    line_position1 = 129
    # Initialize object records
    object_records = {}
    skip_frames = 5
    frame_counter = {}
    cnt = 0
    while True:
        if input_type == 'camera':
            success, img_orig = camera.capture()
            im_name = str(cnt)
        elif input_type == 'video':
            pass
        elif input_type == 'image':
            if cnt >= len(im_paths):
                success = False
            else:
                im_path = im_paths[cnt]
                im_name = os.path.basename(im_path)
                img_orig = cv2.imread(im_path)
                success = True
        if not success:
            break
        result_dict[im_name] = {}
        cnt += 1
        h_orig, w_orig = img_orig.shape[:2]
        target_h = 384
        target_w = 640
        h_ratio = h_orig/target_h
        w_ratio = w_orig/target_w

        img = cv2.resize(img_orig, (640, 384))
        h, w, _ = img.shape
        results_track = model_track.track(
            img, persist=True, conf=0.8, iou=0.7, save_txt=False, tracker="botsort.yaml")
        # results_grade = model_grade(img, persist=True, conf=0.6, iou=0.4, tracker="bytetrack.yaml")
        results_grade = model_grade.predict(
            img, show=False, save=False, show_labels=False, show_conf=False, conf=0.5, iou=0.4, save_txt=False)
        results_track = results_track[0]
        results_grade = results_grade[0]
        # print(results)
        # for r in results:
        # boxes = r.boxes  # Boxes object for bbox outputs
        # masks = r.masks  # Masks object for segment masks outputs
        # probs = r.probs  # Class probabilities for classification outputs
        print(results_track.boxes.id)

        if results_track.boxes.id is not None:
            if len(results_track.boxes.id) > 1:
                # print(results_track.boxes.id)
                pass

        if results_track.masks is not None and results_track.boxes.id is not None:
            masks = results_track.masks.data.cpu()
            ids = results_track.boxes.id.cpu().numpy().astype(int)
            # append_to_xlsx(ids, xlsx_file_path)
            for seg, box, id in zip(masks.data.cpu().numpy(), results_track.boxes, ids):
                seg = cv2.resize(seg, (w, h))
                seg1 = seg.astype(np.uint8)
                res = cv2.findContours(
                    seg1, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                # _, contours, _ = res
                contours, _ = res
                contour = contours[0]
                bboxes = results_grade.boxes.data.cpu().numpy()
                # if len(contour) >= 20 and len(bboxes) > 0:
                if len(contour) >= 5 and len(bboxes) > 0:
                    ellipse = cv2.fitEllipse(contour)
                    center, axes, angle = ellipse
                    long_axis = max(axes)
                    short_axis = min(axes)
                    # Compute endpoints of long axis
                    long_axis_angle_rad = np.deg2rad(angle)
                    long_axis_length = long_axis / 2
                    sin_angle = np.sin(long_axis_angle_rad)
                    cos_angle = np.cos(long_axis_angle_rad)
                    pt1 = (int(center[0] - long_axis_length * sin_angle),
                           int(center[1] + long_axis_length * cos_angle))
                    pt2 = (int(center[0] + long_axis_length * sin_angle),
                           int(center[1] - long_axis_length * cos_angle))

                    # Compute endpoints of short axis
                    short_axis_angle_rad = np.deg2rad(angle + 90)
                    short_axis_length = short_axis / 2
                    sin_angle = np.sin(short_axis_angle_rad)
                    cos_angle = np.cos(short_axis_angle_rad)
                    pt3 = (int(center[0] - short_axis_length * sin_angle),
                           int(center[1] + short_axis_length * cos_angle))
                    pt4 = (int(center[0] + short_axis_length * sin_angle),
                           int(center[1] - short_axis_length * cos_angle))

                    # Draw the long and short axes
                    cv2.line(img, pt1, pt2, (0, 255, 0), 2)
                    cv2.line(img, pt3, pt4, (0, 255, 0), 2)

                    # Display the lengths of the long and short axes
                    length = long_axis_length*2*offset
                    width = short_axis_length*2*offset

                    text = f"L: {length:.2f}, W: {width:.2f}"
                    text_size, _ = cv2.getTextSize(
                        text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                    text_position = (
                        int(center[0] - text_size[0] / 2), int(center[1] + long_axis_length + 20))

                    xmin = int(box.data[0][0])
                    ymin = int(box.data[0][1])
                    xmax = int(box.data[0][2])
                    ymax = int(box.data[0][3])
                    box1 = [xmin, ymin, xmax, ymax]

                    xmin_orig = xmin*w_ratio
                    xmax_orig = xmax*w_ratio
                    ymin_orig = ymin*h_ratio
                    ymax_orig = ymax*h_ratio
                    result_dict[im_name][int(id)] = [
                        xmin_orig, ymin_orig, xmax_orig, ymax_orig]

                    # Bs=results_grade.boxes.data.cpu().numpy()
                    box2s = bboxes

                    final_idx = FindClassAndCof(box1, box2s)
                    cls = box2s[final_idx][5]

                    img = overlay(img, seg, colors[int(cls)], 0.4)

                    if xmin >= line_position1:
                        cv2.putText(img, text, text_position,
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

                    cv2.putText(img, f"Id {id}", (xmax-10, ymax-10),
                                cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, )

                    plot_one_box([xmin, ymin, xmax, ymax], img, colors[int(
                        cls)], f'{class_names[int(cls)]} {float(box2s[final_idx][4]):.3}')

                    cv2.line(img, (line_position, 0),
                             (line_position, 384), (0, 100, 255), 2)
                    cv2.line(img, (line_position1, 0),
                             (line_position1, 384), (0, 100, 255), 2)

                    cv2.putText(img, "Record", (line_position, 50),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    cv2.putText(img, "Grade", (line_position1, 50),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                    # if xmax >= line_position and xmin <= line_position:
                    # if xmax >= line_position and xmin <= line_position:
                    if xmin <= line_position:
                        if xmax >= line_position:
                            cv2.putText(img, f"Start record {id}", (xmin-10, int((ymax+ymin)/2)), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                                        (100, 255, 100), 2)
                        # Record object properties
                        if id not in object_records:
                            object_records[id] = []
                            frame_counter[id] = 0
                        if frame_counter[id] % skip_frames == 0:
                            object_records[id].append({
                                # 'frame': cap.get(cv2.CAP_PROP_POS_FRAMES),
                                'frame': cnt,
                                'class': class_names[int(cls)],
                                "confidence": box2s[final_idx][4],
                                'length': length,
                                'width': width
                            })

                    if xmax-20 >= line_position1 and xmin <= line_position1:
                        cv2.putText(img, f"Start grade {id}", (xmin - 10, int(
                            (ymax + ymin) / 2)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (100, 255, 100), 2)
                        lengths = [record.get('length')
                                   for record in object_records[id]]
                        if lengths is None:
                            print("none")
                        widths = [record['width']
                                  for record in object_records[id]]
                        grades = [record['class']
                                  for record in object_records[id]]
                        t_confidences = [record['confidence']
                                         for record in object_records[id]]
                        m_len = sum(lengths)/len(lengths)
                        m_wid = sum(widths)/len(widths)
                        mid_len = np.median(np.array(lengths))
                        mid_wid = np.median(np.array(widths))
                        confidences = [tensor.item()
                                       for tensor in t_confidences]
                        f_grade = combine_measurements(grades, confidences)
                        #print((id, m_len, mid_len, m_wid, mid_wid, f_grade))
                        F_list=(id, m_len, mid_len, m_wid, mid_wid, f_grade)
                        F_SampleGrade= GetFinalGrade(F_list)
                        print(F_SampleGrade)
                        text = f"F_L: {m_len:.2f}, F_W: {m_wid:.2f}"
                        text_size, _ = cv2.getTextSize(
                            text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                        text_position = (
                            int(center[0] - text_size[0] / 2), int(center[1] + long_axis_length + 20))
                        cv2.putText(img, text, text_position,
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                        key = cv2.waitKey(10)
                        if key & 0xFF == ord('s'):
                            save_data(object_records)

        if input_type == 'image':
            im_dir = os.path.dirname(im_path)
            if save_dir is None:
                save_dir = im_dir
            save_name = im_name.split('.')[0]+'_detected.jpg'
            save_path = join(save_dir, save_name)
            cv2.imwrite(save_path, img)

        # img=cv2.resize(img,(1280,768),interpolation=cv2.INTER_CUBIC)
        cv2.imshow('frame', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    with open(join(save_dir, "result_dict.json"), "w") as file:
        json.dump(result_dict, file)
        print("save result_dict")

    if input_type == 'camera':
        camera.end()
    cv2.destroyAllWindows()


def detect_im():
    dir = r'C:\Users\AFSALab\OneDriveBD\Project\SweetPotatoSorting\video_data\gt\1'
    model_paths = [r'D:\BoyangDeng\SweetPotatoSorting\SweetPotatoGUIProject\model\track\best.pt',
                   r'D:\BoyangDeng\SweetPotatoSorting\SweetPotatoGUIProject\model\grade\best.pt']
    # model_paths = [r'C:\Users\E-ITX\Desktop\BoyangDeng\code\model\track\best.pt',
    #             r'C:\Users\E-ITX\Desktop\BoyangDeng\code\model\grade\best.pt']
    im_paths = list_dir(dir, [], '.jpg')
    save_dir = r'D:\test'
    pipeline('image', im_paths, model_paths, save_dir)


def detect_camera():
    camera = Camera()
    # model_track = YOLO(r"C:\Users\E-ITX\Desktop\BoyangDeng\code\model\train3\best.pt")
    # model_grade = YOLO(r"C:\Users\E-ITX\Desktop\BoyangDeng\code\model\train4\best.pt")
    model_paths = [r'C:\Users\E-ITX\Desktop\BoyangDeng\code\model\track\best.pt',
                   r'C:\Users\E-ITX\Desktop\BoyangDeng\code\model\grade\best.pt']
    pipeline('camera', camera, model_paths)


def main():
    detect_camera()
    # detect_im()


if __name__ == '__main__':
    main()
